"""
Comprehensive XLSX report export for superadmin.
Includes:
- Table states (per-seat totals)
- Chip purchase chronology
- Staff working hours and salary calculations
- Profit/expense summary
"""
from __future__ import annotations

import datetime as dt
import io
from copy import copy
from typing import Any, cast
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session as DBSession, joinedload
from sqlalchemy import func

from ..core.deps import get_current_user, get_db, get_owner_id_for_filter, require_roles
from ..models.db import CasinoBalanceAdjustment, ChipPurchase, DealerRakeEntry, Seat, SeatNameChange, Session, SessionDealerAssignment, Table, User, ChipOp

router = APIRouter(prefix="/api/admin", tags=["admin"])


def _get_working_day_boundaries(date: dt.date) -> tuple[dt.datetime, dt.datetime]:
    """
    Get working day boundaries for a given calendar date.
    Working day: 18:00 (6 PM) to 18:00 (6 PM) of next day.

    This is a full 24-hour window that covers:
    - 18:00-20:00: Pre-session prep time
    - 20:00-18:00 next day: Active playing time

    Args:
        date: Calendar date (YYYY-MM-DD)

    Returns:
        Tuple of (start_datetime, end_datetime) in UTC
    """
    start = dt.datetime.combine(date, dt.time(18, 0, 0))
    end = dt.datetime.combine(date + dt.timedelta(days=1), dt.time(18, 0, 0))
    return start, end


@router.get("/day-summary/preselected-date")
def get_preselected_date(
    db: DBSession = Depends(get_db),
    current_user: User = Depends(require_roles("superadmin", "table_admin")),
):
    """
    Get the preselected date for the daily summary page.

    Returns the starting day of:
    1. Current working day if it's not yet finished (has open sessions)
    2. Most recent working day if current one is finished but next hasn't started

    Working day: 18:00 (6 PM) to 18:00 (6 PM) of next day.

    For table_admin: only considers sessions for tables they own.
    For superadmin: considers all sessions.
    """
    # Multi-tenancy: get owner_id for filtering
    owner_id = get_owner_id_for_filter(current_user)

    now = dt.datetime.utcnow()

    # Determine current working day
    # Working day starts at 18:00 and ends at 18:00 next day
    # So if current time is before 18:00, we're in the working day that started yesterday
    # If current time is 18:00 or later, we're in the working day that started today
    if now.hour < 18:
        # Before 18:00 - we're in the working day that started yesterday
        working_day_start = now.date() - dt.timedelta(days=1)
    else:
        # 18:00 or later - we're in the working day that started today
        working_day_start = now.date()

    # Get working day boundaries
    start_time, end_time = _get_working_day_boundaries(working_day_start)

    # Check for open sessions in current working day
    # Multi-tenancy: filter sessions by tables owned by this user
    if owner_id is not None:
        # table_admin: filter by sessions on tables they own
        owned_table_ids = db.query(Table.id).filter(Table.owner_id == owner_id).subquery()
        query = (
            db.query(Session)
            .filter(Session.created_at >= start_time, Session.created_at < end_time)
            .filter(Session.table_id.in_(owned_table_ids))
        )
    else:
        # superadmin: all sessions
        query = (
            db.query(Session)
            .filter(Session.created_at >= start_time, Session.created_at < end_time)
        )
    open_sessions = query.filter(Session.status == "open").first()
    
    if open_sessions:
        # Current working day is not finished
        return {"date": working_day_start.isoformat()}
    
    # Check if current working day has any sessions at all
    any_sessions = query.first()
    
    if any_sessions:
        # Current working day is finished (all sessions closed)
        return {"date": working_day_start.isoformat()}
    
    # No sessions in current working day - find most recent working day with sessions
    # Look back up to 7 days
    for days_back in range(1, 8):
        prev_day = working_day_start - dt.timedelta(days=days_back)
        prev_start, prev_end = _get_working_day_boundaries(prev_day)

        # Multi-tenancy: filter sessions by tables owned by this user
        if owner_id is not None:
            prev_sessions = (
                db.query(Session)
                .filter(Session.created_at >= prev_start, Session.created_at < prev_end)
                .filter(Session.table_id.in_(owned_table_ids))
                .first()
            )
        else:
            prev_sessions = (
                db.query(Session)
                .filter(Session.created_at >= prev_start, Session.created_at < prev_end)
                .first()
            )

        if prev_sessions:
            return {"date": prev_day.isoformat()}

    # No sessions found in the last 7 days, return current working day
    return {"date": working_day_start.isoformat()}


@router.get("/day-summary")
def get_day_summary(
    date: str = Query(..., description="Date in YYYY-MM-DD format"),
    table_id: int | None = Query(default=None, description="Optional table_id to filter by specific table"),
    db: DBSession = Depends(get_db),
    current_user: User = Depends(require_roles("superadmin", "table_admin")),
):
    """Get day summary data (profit/loss) as JSON for mobile display.

    Multi-tenancy aware:
    - table_admin: shows data for tables they own, staff they own, balance adjustments they own
    - superadmin: shows all data (or filtered by table_id if provided)
    """
    try:
        d = dt.date.fromisoformat(date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    # DIAGNOSTIC LOGGING: Initialize logger
    import logging
    logger = logging.getLogger(__name__)

    # Multi-tenancy: get owner_id for filtering
    owner_id = get_owner_id_for_filter(current_user)
    is_table_admin = owner_id is not None

    # DIAGNOSTIC LOGGING: User role and multi-tenancy
    logger.info(f"=== DAY SUMMARY DIAGNOSTICS FOR {date} ===")
    logger.info(f"--- USER ROLE DIAGNOSTICS ---")
    logger.info(f"current_user.role: {current_user.role}")
    logger.info(f"owner_id: {owner_id}")
    logger.info(f"is_table_admin: {is_table_admin}")

    # Get working day boundaries (18:00 to 18:00 next day)
    start_time, end_time = _get_working_day_boundaries(d)

    # DIAGNOSTIC LOGGING
    logger.info(f"Working day boundaries: {start_time.isoformat()} to {end_time.isoformat()}")

    # Fetch sessions for the working day
    sessions_query = (
        db.query(Session)
        .options(
            joinedload(Session.dealer),
            joinedload(Session.waiter),
            joinedload(Session.dealer_assignments).joinedload(SessionDealerAssignment.dealer),
            joinedload(Session.dealer_assignments).joinedload(SessionDealerAssignment.rake_entries),
        )
        .filter(Session.created_at >= start_time, Session.created_at < end_time)
    )

    # Multi-tenancy: filter by tables owned by user
    if owner_id is not None:
        owned_table_ids = db.query(Table.id).filter(Table.owner_id == owner_id).subquery()
        sessions_query = sessions_query.filter(Session.table_id.in_(owned_table_ids))

    # Additional filter by specific table_id if provided
    if table_id is not None:
        # Verify table ownership for table_admin
        if owner_id is not None:
            table = db.query(Table).filter(Table.id == table_id, Table.owner_id == owner_id).first()
            if not table:
                raise HTTPException(status_code=403, detail="Forbidden for this table")
        sessions_query = sessions_query.filter(Session.table_id == table_id)

    sessions = sessions_query.order_by(Session.table_id.asc(), Session.created_at.asc()).all()

    session_ids = [cast(str, s.id) for s in sessions]

    # Fetch seats for all sessions
    seats_by_session: dict[str, list[Seat]] = {}
    if session_ids:
        seats = (
            db.query(Seat)
            .filter(Seat.session_id.in_(session_ids))
            .all()
        )
        for seat in seats:
            sid = cast(str, seat.session_id)
            seats_by_session.setdefault(sid, []).append(seat)

    # Fetch all chip purchases for the date
    purchases = (
        db.query(ChipPurchase)
        .filter(ChipPurchase.session_id.in_(session_ids))
        .all()
    ) if session_ids else []

    # Fetch balance adjustments for the working day (multi-tenancy filtered)
    balance_query = (
        db.query(CasinoBalanceAdjustment)
        .options(joinedload(CasinoBalanceAdjustment.created_by))
        .filter(CasinoBalanceAdjustment.created_at >= start_time, CasinoBalanceAdjustment.created_at < end_time)
    )
    if owner_id is not None:
        balance_query = balance_query.filter(CasinoBalanceAdjustment.owner_id == owner_id)
    balance_adjustments = balance_query.order_by(CasinoBalanceAdjustment.created_at.asc()).all()

    # DIAGNOSTIC LOGGING: Balance adjustments
    logger.info(f"--- BALANCE ADJUSTMENTS DIAGNOSTICS ---")
    logger.info(f"balance_adjustments length: {len(balance_adjustments)}")
    logger.info(f"is_table_admin: {is_table_admin}")

    # Fetch staff (multi-tenancy filtered)
    # For table_admin: only staff they own
    # For superadmin: all staff
    staff_query = db.query(User).filter(User.role.in_(["dealer", "waiter"]))
    if owner_id is not None:
        staff_query = staff_query.filter(User.owner_id == owner_id)
    staff = staff_query.all()

    # DIAGNOSTIC LOGGING: Staff
    logger.info(f"--- STAFF DIAGNOSTICS ---")
    logger.info(f"staff length: {len(staff)}")
    logger.info(f"is_table_admin: {is_table_admin}")

    # Calculate totals
    total_chip_income_cash = 0  # Cash buyins (positive only)
    total_chip_cashout = 0  # Cash cashouts (absolute value, negative amounts)
    total_chip_income_credit = 0  # Credit buyins (informational only, NOT in rake calculation)
    total_balance_adjustments_profit = 0  # Positive adjustments
    total_balance_adjustments_expense = 0  # Negative adjustments (absolute value)

    for p in purchases:
        amount = int(cast(int, p.amount))
        if amount > 0:  # Buyin
            if p.payment_type == "credit":
                total_chip_income_credit += amount  # Track for info only
            else:
                total_chip_income_cash += amount
        elif amount < 0:  # Cashout (negative amount = expense)
            # Cashouts are always cash payments back to players
            total_chip_cashout += abs(amount)  # Track cashouts separately

    # Process balance adjustments
    balance_adjustments_list = []
    for adj in balance_adjustments:
        amount = int(cast(int, adj.amount))
        created_by_username = cast(str, adj.created_by.username) if adj.created_by else "—"
        
        adjustment_data = {
            "id": int(cast(int, adj.id)),
            "created_at": cast(dt.datetime, adj.created_at).isoformat(),
            "amount": amount,
            "comment": cast(str, adj.comment),
            "created_by_username": created_by_username,
        }
        balance_adjustments_list.append(adjustment_data)
        
        if amount > 0:
            total_balance_adjustments_profit += amount
        else:
            total_balance_adjustments_expense += abs(amount)

    # Calculate staff salary for both superadmin and table_admin
    # For table_admin: only staff who worked on their table
    # For superadmin: all staff (or filtered by table_id if provided)
    total_salary = 0
    staff_details = []
    for person in staff:
        role = cast(str, person.role)
        hourly_rate = int(cast(int, person.hourly_rate)) if person.hourly_rate else 0

        if role == "dealer":
            hours = _calculate_dealer_hours(sessions, int(cast(int, person.id)))
        else:
            hours = _calculate_waiter_hours(sessions, int(cast(int, person.id)))

        salary = round(hours * hourly_rate)
        if hours > 0:
            staff_details.append({
                "name": person.username,
                "role": role,
                "hours": round(hours, 2),
                "hourly_rate": hourly_rate,
                "salary": salary,
            })
        total_salary += salary

    # DIAGNOSTIC LOGGING: Total salary
    logger.info(f"--- SALARY DIAGNOSTICS ---")
    logger.info(f"total_salary: {total_salary}")
    logger.info(f"is_table_admin: {is_table_admin}")

    # Calculate net per-seat totals
    total_player_balance = 0
    for sid, seats in seats_by_session.items():
        for seat in seats:
            total_player_balance += int(cast(int, seat.total))

    # Gross rake ("грязный") = sum of manually entered rake entries from all dealer assignments
    # Rake entries are added by table admin during the session via the "Update Rake" feature
    gross_rake = 0
    for s in sessions:
        for assignment in s.dealer_assignments:
            for entry in (assignment.rake_entries or []):
                gross_rake += int(cast(int, entry.amount))

    # Net result for the day = gross_rake - salaries + balance_adjustments_profit - balance_adjustments_expense
    net_result = gross_rake - total_salary + total_balance_adjustments_profit - total_balance_adjustments_expense

    # DIAGNOSTIC LOGGING
    logger.info(f"--- CALCULATION COMPONENTS ---")
    logger.info(f"total_chip_income_cash (cash buyins): {total_chip_income_cash}")
    logger.info(f"total_chip_cashout (cashouts to players): {total_chip_cashout}")
    logger.info(f"total_player_balance (chips players have): {total_player_balance}")
    logger.info(f"total_salary: {total_salary}")
    logger.info(f"total_chip_income_credit (credit buyins - info only): {total_chip_income_credit}")
    logger.info(f"total_balance_adjustments_profit: {total_balance_adjustments_profit}")
    logger.info(f"total_balance_adjustments_expense: {total_balance_adjustments_expense}")
    logger.info(f"--- FORMULAS ---")
    logger.info(f"gross_rake = sum of dealer assignment rakes = {gross_rake}")
    logger.info(
        f"net_result = {gross_rake} - {total_salary} + {total_balance_adjustments_profit} - {total_balance_adjustments_expense} = {net_result}"
    )
    logger.info(f"--- BALANCE ADJUSTMENTS DETAIL ---")
    for adj in balance_adjustments_list:
        logger.info(f"  ID {adj['id']}: {adj['comment']} = {adj['amount']} ₪ (by {adj['created_by_username']})")
    logger.info(f"--- SESSIONS DETAIL ---")
    for s in sessions:
        sid = cast(str, s.id)
        seats = seats_by_session.get(sid, [])
        session_balance = sum(int(cast(int, seat.total)) for seat in seats)
        session_rake = sum(int(cast(int, a.rake)) for a in s.dealer_assignments if a.rake is not None)
        logger.info(f"  Session {sid} (Table {s.table_id}, {s.status}): player_balance = {session_balance}, rake = {session_rake}")
    logger.info(f"=== END DIAGNOSTICS ===")

    open_sessions = len([s for s in sessions if s.status == "open"])

    # DIAGNOSTIC LOGGING: Complete response structure
    # Build response - both superadmin and table_admin now see all fields
    # For table_admin: data is filtered to their table only
    # For superadmin: data is casino-wide or filtered by table_id if provided
    response_data = {
        "date": date,
        "working_day_start": start_time.isoformat(),
        "working_day_end": end_time.isoformat(),
        "income": {
            "gross_rake": gross_rake,
            "balance_adjustments": total_balance_adjustments_profit,
        },
        "expenses": {
            "salaries": total_salary,
            "balance_adjustments": total_balance_adjustments_expense,
        },
        "result": net_result,
        "info": {
            "buyin_cash": total_chip_income_cash,
            "buyin_credit": total_chip_income_credit,  # Informational only
            "cashout": total_chip_cashout,
            "player_balance": total_player_balance,
            "total_sessions": len(sessions),
            "open_sessions": open_sessions,
        },
        "staff": staff_details,
        "balance_adjustments": balance_adjustments_list,
    }
    logger.info(f"=== COMPLETE RESPONSE STRUCTURE ===")
    logger.info(f"Response: {response_data}")
    logger.info(f"=== END RESPONSE DIAGNOSTICS ===")

    return response_data


# Style constants
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
MONEY_POSITIVE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
MONEY_NEGATIVE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
# Dark shades for payment types
CASH_DARK_FILL = PatternFill(start_color="006400", end_color="006400", fill_type="solid")  # Dark green
CREDIT_DARK_FILL = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")  # Dark red

# Player block colored cells (based on user's color scheme)
CHIPS_TAKEN_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange - Жетоны(-)
CHIPS_RETURNED_FILL = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")  # Cyan - Жетоны(+)
RESULT_FILL = PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type="solid")  # Magenta - Результат
CASH_GIVEN_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow - Наличных

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)




def _style_header(ws, row: int, cols: int):
    """Apply header styling to a row."""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _auto_width(ws):
    """Auto-adjust column widths with better minimum width for readability."""
    from openpyxl.cell.cell import MergedCell

    for column_cells in ws.columns:
        max_length = 0
        column = None

        # Find the first non-merged cell to get the column letter
        for cell in column_cells:
            if not isinstance(cell, MergedCell):
                column = cell.column_letter
                break

        if column is None:
            continue  # Skip if all cells in column are merged

        for cell in column_cells:
            try:
                if not isinstance(cell, MergedCell) and cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        # Increased minimum width from (max_length + 2) to (max_length + 4) for better fit
        # Also increased maximum from 50 to 60 for longer text fields
        adjusted_width = min(max_length + 4, 60)
        # Ensure minimum width of 12 for very short columns
        adjusted_width = max(adjusted_width, 12)
        ws.column_dimensions[column].width = adjusted_width


def _apply_session_border(ws, start_row: int, end_row: int, max_col: int = 5):
    """Apply a border around a session block to visually separate it."""
    thin_border_side = Side(style='thin', color='808080')

    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)

            # Determine which borders to apply
            top = thin_border_side if row == start_row else None
            bottom = thin_border_side if row == end_row else None
            left = thin_border_side if col == 1 else None
            right = thin_border_side if col == max_col else None

            # Only update border if we need to add one
            if top or bottom or left or right:
                current_border = cell.border
                cell.border = Border(
                    left=left or current_border.left,
                    right=right or current_border.right,
                    top=top or current_border.top,
                    bottom=bottom or current_border.bottom
                )


def _calculate_waiter_hours(
    sessions: list[Session],
    waiter_id: int,
) -> float:
    """
    Calculate waiter working hours accounting for overlapping sessions.
    Waiters can work on multiple sessions at a time, so we need to merge
    overlapping time intervals.
    """
    intervals: list[tuple[dt.datetime, dt.datetime]] = []

    for s in sessions:
        if s.waiter_id != waiter_id:
            continue
        start = cast(dt.datetime, s.created_at)
        end = cast(dt.datetime, s.closed_at) if s.closed_at else dt.datetime.utcnow()
        intervals.append((start, end))

    if not intervals:
        return 0.0

    # Sort by start time
    intervals.sort(key=lambda x: x[0])

    # Merge overlapping intervals
    merged: list[tuple[dt.datetime, dt.datetime]] = [intervals[0]]
    for start, end in intervals[1:]:
        last_start, last_end = merged[-1]
        if start <= last_end:
            # Overlapping, extend the end if needed
            merged[-1] = (last_start, max(last_end, end))
        else:
            merged.append((start, end))

    # Sum total hours
    total_seconds = sum((end - start).total_seconds() for start, end in merged)
    return total_seconds / 3600.0


def _get_waiter_time_range(
    sessions: list[Session],
    waiter_id: int,
) -> tuple[dt.datetime | None, dt.datetime | None]:
    """
    Get the overall time range for a waiter's work period.
    Returns (earliest_start, latest_end) or (None, None) if no sessions.
    """
    intervals: list[tuple[dt.datetime, dt.datetime]] = []

    for s in sessions:
        if s.waiter_id != waiter_id:
            continue
        start = cast(dt.datetime, s.created_at)
        end = cast(dt.datetime, s.closed_at) if s.closed_at else dt.datetime.utcnow()
        intervals.append((start, end))

    if not intervals:
        return None, None

    earliest_start = min(start for start, _ in intervals)
    latest_end = max(end for _, end in intervals)
    return earliest_start, latest_end


def _calculate_dealer_hours(
    sessions: list[Session],
    dealer_id: int,
) -> float:
    """
    Calculate dealer working hours using SessionDealerAssignment records.
    This accounts for dealer changes within a session.

    If a session has dealer_assignments, use those for accurate time tracking.
    Otherwise, fall back to the legacy dealer_id field for backward compatibility.
    """
    total_seconds = 0.0
    now = dt.datetime.utcnow()

    for s in sessions:
        # Check if session has dealer assignments (new method)
        if s.dealer_assignments:
            for assignment in s.dealer_assignments:
                if int(cast(int, assignment.dealer_id)) != dealer_id:
                    continue
                start = cast(dt.datetime, assignment.started_at)
                end = cast(dt.datetime, assignment.ended_at) if assignment.ended_at else now
                total_seconds += (end - start).total_seconds()
        else:
            # Fallback to legacy method for sessions without dealer_assignments
            if s.dealer_id != dealer_id:
                continue
            start = cast(dt.datetime, s.created_at)
            end = cast(dt.datetime, s.closed_at) if s.closed_at else now
            total_seconds += (end - start).total_seconds()

    return total_seconds / 3600.0


def _calculate_session_dealer_earnings(
    session: Session,
    db,
) -> list[dict]:
    """
    Calculate earnings for all dealers who worked on this session.
    Returns list of dicts with dealer info, hours worked, and salary.
    """
    from ..models.db import SessionDealerAssignment, User

    earnings = []
    now = dt.datetime.utcnow()

    # Get all dealer assignments for this session
    if session.dealer_assignments:
        # Use dealer assignments (new method)
        for assignment in session.dealer_assignments:
            dealer = assignment.dealer
            if not dealer:
                continue

            start = cast(dt.datetime, assignment.started_at)
            end = cast(dt.datetime, assignment.ended_at) if assignment.ended_at else (
                cast(dt.datetime, session.closed_at) if session.closed_at else now
            )
            hours = (end - start).total_seconds() / 3600.0
            hourly_rate = int(cast(int, dealer.hourly_rate)) if dealer.hourly_rate else 0
            salary = round(hours * hourly_rate)

            # Sum rake entries for this assignment
            rake = sum(int(cast(int, entry.amount)) for entry in (assignment.rake_entries or []))

            earnings.append({
                "dealer_name": cast(str, dealer.username),
                "hours": hours,
                "hourly_rate": hourly_rate,
                "salary": salary,
                "rake": rake,
            })
    elif session.dealer_id:
        # Fallback to legacy method for sessions without dealer_assignments
        dealer = db.query(User).filter(User.id == session.dealer_id).first()
        if dealer:
            start = cast(dt.datetime, session.created_at)
            end = cast(dt.datetime, session.closed_at) if session.closed_at else now
            hours = (end - start).total_seconds() / 3600.0
            hourly_rate = int(cast(int, dealer.hourly_rate)) if dealer.hourly_rate else 0
            salary = round(hours * hourly_rate)

            earnings.append({
                "dealer_name": cast(str, dealer.username),
                "hours": hours,
                "hourly_rate": hourly_rate,
                "salary": salary,
                "rake": 0,  # Legacy sessions don't have per-dealer rake
            })

    return earnings


def _calculate_session_waiter_earnings(
    session: Session,
    db,
) -> dict | None:
    """
    Calculate earnings for the waiter who worked on this session.
    Returns dict with waiter info, hours worked, and salary, or None if no waiter.
    """
    from ..models.db import User

    if not session.waiter_id:
        return None

    waiter = db.query(User).filter(User.id == session.waiter_id).first()
    if not waiter:
        return None

    now = dt.datetime.utcnow()
    start = cast(dt.datetime, session.created_at)
    end = cast(dt.datetime, session.closed_at) if session.closed_at else now
    hours = (end - start).total_seconds() / 3600.0
    hourly_rate = int(cast(int, waiter.hourly_rate)) if waiter.hourly_rate else 0
    salary = round(hours * hourly_rate)

    return {
        "waiter_name": cast(str, waiter.username),
        "hours": hours,
        "hourly_rate": hourly_rate,
        "salary": salary,
    }


@router.get("/export-report")
def export_report(
    date: str = Query(..., description="YYYY-MM-DD"),
    table_id: int | None = Query(default=None, description="Optional table_id to filter by specific table"),
    db: DBSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Generate comprehensive XLSX report for a specific date.

    Multi-tenancy aware:
    - table_admin: includes data for tables/staff/balance adjustments they own
    - superadmin: includes all data (or filtered by table_id if provided)
    """
    # Check user role
    role = cast(str, user.role)
    if role not in ("superadmin", "table_admin"):
        raise HTTPException(status_code=403, detail="Forbidden")

    # Multi-tenancy: get owner_id for filtering
    owner_id = get_owner_id_for_filter(user)
    is_table_admin = owner_id is not None

    try:
        d = dt.date.fromisoformat(date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format (expected YYYY-MM-DD)")

    # Get working day boundaries (18:00 to 18:00 next day)
    start_time, end_time = _get_working_day_boundaries(d)

    # Fetch tables (multi-tenancy filtered)
    tables_query = db.query(Table)
    if owner_id is not None:
        tables_query = tables_query.filter(Table.owner_id == owner_id)
    tables = tables_query.order_by(Table.id.asc()).all()

    # Fetch sessions for the working day
    sessions_query = (
        db.query(Session)
        .options(
            joinedload(Session.dealer),
            joinedload(Session.waiter),
            joinedload(Session.dealer_assignments).joinedload(SessionDealerAssignment.dealer),
            joinedload(Session.dealer_assignments).joinedload(SessionDealerAssignment.rake_entries),
        )
        .filter(Session.created_at >= start_time, Session.created_at < end_time)
    )

    # Multi-tenancy: filter by tables owned by user
    if owner_id is not None:
        owned_table_ids = db.query(Table.id).filter(Table.owner_id == owner_id).subquery()
        sessions_query = sessions_query.filter(Session.table_id.in_(owned_table_ids))

    # Additional filter by specific table_id if provided
    if table_id is not None:
        # Verify table ownership for table_admin
        if owner_id is not None:
            table = db.query(Table).filter(Table.id == table_id, Table.owner_id == owner_id).first()
            if not table:
                raise HTTPException(status_code=403, detail="Forbidden for this table")
        sessions_query = sessions_query.filter(Session.table_id == table_id)

    sessions = sessions_query.order_by(Session.table_id.asc(), Session.created_at.asc()).all()

    session_ids = [cast(str, s.id) for s in sessions]

    # Fetch seats for all sessions
    seats_by_session: dict[str, list[Seat]] = {}
    if session_ids:
        seats = (
            db.query(Seat)
            .filter(Seat.session_id.in_(session_ids))
            .order_by(Seat.session_id.asc(), Seat.seat_no.asc())
            .all()
        )
        for seat in seats:
            sid = cast(str, seat.session_id)
            seats_by_session.setdefault(sid, []).append(seat)

    # Fetch all chip purchases for the date (used for payment_type lookup)
    purchases = (
        db.query(ChipPurchase)
        .options(joinedload(ChipPurchase.created_by))
        .filter(ChipPurchase.session_id.in_(session_ids))
        .order_by(ChipPurchase.created_at.asc())
        .all()
    ) if session_ids else []

    # Fetch all chip operations for the date (primary source of truth - never deleted)
    chip_ops = (
        db.query(ChipOp)
        .filter(ChipOp.session_id.in_(session_ids))
        .order_by(ChipOp.created_at.asc())
        .all()
    ) if session_ids else []

    # Fetch balance adjustments for the working day (multi-tenancy filtered)
    balance_query = (
        db.query(CasinoBalanceAdjustment)
        .options(joinedload(CasinoBalanceAdjustment.created_by))
        .filter(CasinoBalanceAdjustment.created_at >= start_time, CasinoBalanceAdjustment.created_at < end_time)
    )
    if owner_id is not None:
        balance_query = balance_query.filter(CasinoBalanceAdjustment.owner_id == owner_id)
    balance_adjustments = balance_query.order_by(CasinoBalanceAdjustment.created_at.asc()).all()

    # Fetch staff (multi-tenancy filtered)
    staff_query = db.query(User).filter(User.role.in_(["dealer", "waiter"]))
    if owner_id is not None:
        staff_query = staff_query.filter(User.owner_id == owner_id)
    staff = staff_query.all()

    # Load template.xlsx and fill it with data
    import os
    template_path = os.path.join(os.path.dirname(__file__), "..", "..", "template.xlsx")
    if not os.path.exists(template_path):
        # Fallback for Docker container
        template_path = "/app/template.xlsx"

    wb = load_workbook(template_path)
    ws = wb.active

    # Fill template with data
    _fill_template_with_data(ws, sessions, seats_by_session, purchases, chip_ops, balance_adjustments, staff, d, db)

    # Generate file
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"casino_report_{date}.xlsx"

    headers = {
        "Content-Disposition": (
            f'attachment; filename="{filename}"; '
            f"filename*=UTF-8''{quote(filename)}"
        )
    }

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


def _fill_template_with_data(
    ws,
    sessions: list[Session],
    seats_by_session: dict[str, list[Seat]],
    purchases: list[ChipPurchase],
    chip_ops: list[ChipOp],
    balance_adjustments: list[CasinoBalanceAdjustment],
    staff: list[User],
    report_date: dt.date,
    db: DBSession,
):
    """
    Fill the template.xlsx with actual data.

    Template structure (dynamic rows for seats):
    - Row 1: Seat number headers (we fill № 1 through № 10 in columns C,E,G,I,K,M,O,Q,S,U)
    - Rows 2-N: Player names + chip purchases (dynamically sized based on data)
    - Rows N+1 to N+6: Summary rows (Жетоны(-), Жетоны(+), Результат, etc.)

    After seat section, staff section follows with adjusted row numbers.
    """
    if not sessions:
        return  # Leave template as-is if no data

    from openpyxl.styles import PatternFill
    import re

    no_fill = PatternFill(fill_type=None)
    MAX_SEATS = 10
    TEMPLATE_DATA_ROWS = 14  # Template has rows 2-15 for data (14 rows)
    TEMPLATE_SUMMARY_START_ROW = 16  # Original summary rows in template

    # Define consistent font for seats area (matching template style)
    seats_font = Font(name="Roboto Mono", size=11)
    seats_font_bold = Font(name="Roboto Mono", size=11, bold=True)

    # Read background fills from row 1 for seat columns (yellow for value, blue for time)
    # We'll read from seat 1's columns (B=2 for value, C=3 for time) and apply to all seats
    seat_value_fill = copy(ws.cell(row=1, column=2).fill)  # Yellow from column B
    seat_time_fill = copy(ws.cell(row=1, column=3).fill)   # Blue from column C

    # --- Step 1: Collect all data for each seat ---
    # For each seat, we need to interleave player names (including changes) and chip purchases

    # Get all session IDs for the report
    session_ids = [cast(str, s.id) for s in sessions]

    # Query all SeatNameChange records for these sessions
    name_changes = db.query(SeatNameChange).filter(
        SeatNameChange.session_id.in_(session_ids)
    ).all()

    # Group name changes by seat
    name_changes_by_seat: dict[int, list[SeatNameChange]] = {}
    for nc in name_changes:
        seat_no = int(cast(int, nc.seat_no))
        if seat_no not in name_changes_by_seat:
            name_changes_by_seat[seat_no] = []
        name_changes_by_seat[seat_no].append(nc)

    # Sort name changes by time
    for seat_no in name_changes_by_seat:
        name_changes_by_seat[seat_no].sort(key=lambda nc: cast(dt.datetime, nc.created_at))

    # Build a map of chip_op_id to purchase for payment_type lookup
    purchase_by_op_id: dict[int, ChipPurchase] = {}
    for p in purchases:
        op_id = int(cast(int, p.chip_op_id))
        purchase_by_op_id[op_id] = p

    # Collect all chip operations grouped by seat (chip_ops are never deleted, unlike chip_purchases)
    chip_ops_by_seat: dict[int, list[ChipOp]] = {}
    for op in chip_ops:
        seat_no = int(cast(int, op.seat_no))
        if seat_no not in chip_ops_by_seat:
            chip_ops_by_seat[seat_no] = []
        chip_ops_by_seat[seat_no].append(op)

    # Sort chip operations by time
    for seat_no in chip_ops_by_seat:
        chip_ops_by_seat[seat_no].sort(key=lambda op: cast(dt.datetime, op.created_at))

    # Get initial player names by seat (from Seat records)
    initial_player_by_seat: dict[int, tuple[str, dt.datetime]] = {}
    for session in sessions:
        session_id = cast(str, session.id)
        session_created = cast(dt.datetime, session.created_at)
        seats = seats_by_session.get(session_id, [])
        for seat in seats:
            seat_no = int(cast(int, seat.seat_no))
            # We need to find the initial player name (first name before any changes)
            # If there are name changes, the first change's old_name is the initial
            seat_name_changes = name_changes_by_seat.get(seat_no, [])
            if seat_name_changes:
                initial_name = seat_name_changes[0].old_name
            else:
                # No changes, use current name
                initial_name = seat.player_name
            if initial_name and seat_no not in initial_player_by_seat:
                initial_player_by_seat[seat_no] = (cast(str, initial_name), session_created)

    # Build player blocks for each seat
    # A player block = all events from when player sits down until they leave
    # Each block will have: player name, chip operations, then summary rows
    seat_player_blocks: dict[int, list[dict]] = {}

    for seat_no in range(1, MAX_SEATS + 1):
        blocks = []
        current_block = None

        # Collect all events for this seat
        all_events: list[tuple[dt.datetime, str, Any]] = []

        # Add initial player name if exists
        if seat_no in initial_player_by_seat:
            name, ts = initial_player_by_seat[seat_no]
            all_events.append((ts, "player", name))

        # Add name changes
        for nc in name_changes_by_seat.get(seat_no, []):
            ts = cast(dt.datetime, nc.created_at)
            change_type = nc.change_type or "name_change"
            if change_type == "player_left":
                all_events.append((ts, "player_left", nc.old_name))
            elif nc.new_name:
                all_events.append((ts, "player", nc.new_name))

        # Add chip operations (using chip_ops instead of purchases for complete history)
        for op in chip_ops_by_seat.get(seat_no, []):
            ts = cast(dt.datetime, op.created_at)
            all_events.append((ts, "chip_op", op))

        # Sort all events by timestamp
        all_events.sort(key=lambda e: e[0])

        # Group events into player blocks
        # Key: player_name -> block dict
        player_block_map: dict[str, dict] = {}

        for ts, event_type, data in all_events:
            if event_type == "player":
                player_name = data
                # Check if this player already has a block (re-entry to same seat)
                if player_name in player_block_map:
                    # Reopen existing block
                    current_block = player_block_map[player_name]
                    current_block["events"].append((ts, "player", player_name))
                else:
                    # Start new block for this player
                    current_block = {
                        "player_name": player_name,
                        "start_time": ts,
                        "events": [(ts, "player", player_name)],
                        "chip_ops": [],
                    }
                    player_block_map[player_name] = current_block
                    blocks.append(current_block)
            elif event_type == "player_left":
                # Player leaves - add to current block
                # Don't remove from map (they might come back)
                if current_block is not None:
                    current_block["events"].append((ts, "player_left", data))
                    current_block["end_time"] = ts
                    current_block = None  # Clear current but keep in map
            elif event_type == "chip_op":
                # Chip operation - add to current block
                if current_block is not None:
                    op = data  # This is a ChipOp object
                    current_block["events"].append((ts, "chip_op", op))
                    current_block["chip_ops"].append(op)

        seat_player_blocks[seat_no] = blocks

    # --- Step 2: Calculate max data rows needed ---
    # For each seat, count: events (excluding player_left) + 4 summary rows per player block
    max_data_rows = 0
    for seat_no in range(1, MAX_SEATS + 1):
        blocks = seat_player_blocks.get(seat_no, [])
        total_rows = 0
        for block in blocks:
            # Count only player and chip_op events (skip player_left)
            event_count = sum(1 for _, et, _ in block["events"] if et in ("player", "chip_op"))
            total_rows += event_count
            total_rows += 4  # 4 summary rows per block
        max_data_rows = max(max_data_rows, total_rows)

    # Ensure at least TEMPLATE_DATA_ROWS (for when there's little data)
    max_data_rows = max(max_data_rows, TEMPLATE_DATA_ROWS)

    # --- Step 3: Insert rows if needed ---
    rows_to_insert = max_data_rows - TEMPLATE_DATA_ROWS
    if rows_to_insert > 0:
        # Insert rows before the summary section (before row 16)
        ws.insert_rows(TEMPLATE_SUMMARY_START_ROW, rows_to_insert)

    # Calculate new summary row positions (no longer used for per-seat summaries)
    summary_start_row = TEMPLATE_SUMMARY_START_ROW + rows_to_insert  # 16 + inserted rows
    data_end_row = summary_start_row - 1  # Last data row

    # --- Step 4: Update seat headers ---
    for seat_no in range(1, MAX_SEATS + 1):
        col_time = 3 + (seat_no - 1) * 2  # C, E, G, I, K, M, O, Q, S, U
        cell = ws.cell(row=1, column=col_time, value=f"№ {seat_no}")
        cell.font = seats_font_bold

    # --- Step 5: Clear unused seat columns (seats 11-12 in template) and column V ---
    # Clear columns V (22) through Y (25) to remove any leftover borders/data
    for col in range(22, 26):
        for row in range(1, summary_start_row + 15):
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.fill = no_fill
            cell.border = Border()

    # --- Step 6: Clear ALL seat data area AND old summary rows ---
    # Clear from row 2 to summary_start_row + 10 (to catch old summary rows)
    clear_end_row = summary_start_row + 10
    for row in range(2, clear_end_row):
        for col in range(2, 22):  # B to U (all seat columns)
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.fill = no_fill
            cell.border = Border()

    # Now re-fill with our player block data
    actual_max_row = 1  # Track actual max row used
    for seat_no in range(1, MAX_SEATS + 1):
        col_value = 2 + (seat_no - 1) * 2
        col_time = col_value + 1
        blocks = seat_player_blocks.get(seat_no, [])
        current_row = 2

        for block in blocks:
            block_start_row = current_row
            player_entry_count = 0  # Track re-entries within the same block

            # Fill events (skip player_left events - borders are enough separation)
            for ts, event_type, data in block["events"]:
                if event_type == "player":
                    player_entry_count += 1
                    cell_val = ws.cell(row=current_row, column=col_value, value=data)
                    cell_val.font = seats_font
                    cell_val.fill = seat_value_fill  # Yellow background
                    cell_time = ws.cell(row=current_row, column=col_time, value=ts.strftime("%H:%M"))
                    cell_time.font = seats_font
                    cell_time.fill = seat_time_fill  # Blue background
                    # Add top border for re-entries (not the first entry)
                    if player_entry_count > 1:
                        thick_side = Side(style='medium')
                        for c in [col_value, col_time]:
                            cell = ws.cell(row=current_row, column=c)
                            cell.border = Border(top=thick_side)
                    current_row += 1
                elif event_type == "chip_op":
                    op = data
                    amount = int(cast(int, op.amount))
                    cell_val = ws.cell(row=current_row, column=col_value, value=amount)
                    cell_val.font = seats_font
                    cell_val.fill = seat_value_fill  # Yellow background
                    cell_time = ws.cell(row=current_row, column=col_time, value=ts.strftime("%H:%M"))
                    cell_time.font = seats_font
                    cell_time.fill = seat_time_fill  # Blue background
                    current_row += 1
                # Skip player_left events - the block border provides visual separation

            # Calculate summaries using chip_ops
            chips_taken = 0
            chips_returned = 0
            cash_given = 0

            for op in block["chip_ops"]:
                op_id = int(cast(int, op.id))
                amount = int(cast(int, op.amount))
                # Look up payment_type from ChipPurchase (if exists)
                purchase = purchase_by_op_id.get(op_id)
                payment_type = cast(str, purchase.payment_type) if purchase and purchase.payment_type else "cash"

                if amount > 0:
                    chips_taken += amount
                    if payment_type == "cash":
                        cash_given += amount
                else:
                    chips_returned += abs(amount)

            result = chips_taken - chips_returned

            # Summary rows (use bold font for summary)
            cell = ws.cell(row=current_row, column=col_value, value=chips_taken)
            cell.fill = CHIPS_TAKEN_FILL
            cell.font = seats_font_bold
            cell = ws.cell(row=current_row, column=col_time, value="Жетоны(-)")
            cell.fill = CHIPS_TAKEN_FILL
            cell.font = seats_font_bold
            current_row += 1

            cell = ws.cell(row=current_row, column=col_value, value=chips_returned)
            cell.fill = CHIPS_RETURNED_FILL
            cell.font = seats_font_bold
            cell = ws.cell(row=current_row, column=col_time, value="Жетоны(+)")
            cell.fill = CHIPS_RETURNED_FILL
            cell.font = seats_font_bold
            current_row += 1

            cell = ws.cell(row=current_row, column=col_value, value=result)
            cell.fill = RESULT_FILL
            cell.font = seats_font_bold
            cell = ws.cell(row=current_row, column=col_time, value="Результат")
            cell.fill = RESULT_FILL
            cell.font = seats_font_bold
            current_row += 1

            cell = ws.cell(row=current_row, column=col_value, value=cash_given)
            cell.fill = CASH_GIVEN_FILL
            cell.font = seats_font_bold
            cell = ws.cell(row=current_row, column=col_time, value="Наличных")
            cell.fill = CASH_GIVEN_FILL
            cell.font = seats_font_bold
            current_row += 1

            # Add borders around this player block
            block_end_row = current_row - 1
            thick_side = Side(style='medium')
            for r in range(block_start_row, block_end_row + 1):
                for c in [col_value, col_time]:
                    cell = ws.cell(row=r, column=c)
                    current_border = cell.border
                    left = thick_side if c == col_value else current_border.left
                    right = thick_side if c == col_time else current_border.right
                    top = thick_side if r == block_start_row else current_border.top
                    bottom = thick_side if r == block_end_row else current_border.bottom
                    cell.border = Border(left=left, right=right, top=top, bottom=bottom)

        # Track max row used across all seats
        if current_row > actual_max_row:
            actual_max_row = current_row - 1

    # --- Step 7: Add medium outer border around entire seats area ---
    # Use actual_max_row instead of data_end_row to avoid empty rows
    seats_area_end_row = actual_max_row
    thick_side = Side(style='medium')
    for r in range(1, seats_area_end_row + 1):
        for c in range(2, 22):  # B to U (columns 2-21)
            cell = ws.cell(row=r, column=c)
            current_border = cell.border
            left = thick_side if c == 2 else current_border.left
            right = thick_side if c == 21 else current_border.right
            top = thick_side if r == 1 else current_border.top
            bottom = thick_side if r == seats_area_end_row else current_border.bottom
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    # Clear any rows between actual_max_row and data_end_row (remove empty rows with borders)
    # Include column A (1) to clear any leftover borders between A and B
    for r in range(actual_max_row + 1, data_end_row + 1):
        for c in range(1, 23):  # Columns A through V (1-22)
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.fill = no_fill
            cell.border = Border()

    # --- Step 9: Delete unused rows from original template ---
    # Original template had rows 24-44 for second seat section (21 rows)
    # After inserting rows, these are at (24 + rows_to_insert) to (44 + rows_to_insert)
    rows_24_44_start = 24 + rows_to_insert
    ws.delete_rows(rows_24_44_start, 21)

    # After deleting 21 rows, the dealer section position is:
    # Original row 47 + rows_to_insert - 21 = 26 + rows_to_insert
    # But we need to clear rows between actual_max_row and the dealer section
    # Note: actual_max_row is still valid since it's before the deleted rows

    # Clear any remaining empty rows with borders after deletion
    # The dealer section starts at row 26 + rows_to_insert after the deletion
    DEALER_SECTION_START_ROW = 26 + rows_to_insert
    for r in range(actual_max_row + 1, DEALER_SECTION_START_ROW):
        for c in range(1, 23):  # Columns A through V (1-22)
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.fill = no_fill
            cell.border = Border()
            cell.font = Font()

    # === SECTION: Dealer rake entries ===
    # After row operations: original row 47 -> summary_start_row + 6 + gap (usually row 23)
    # Then after deleting 21 rows, it shifts up by 21
    # Net effect: original 47 -> (47 + rows_to_insert - 21) = 47 - 21 + rows_to_insert = 26 + rows_to_insert
    # But we need to account for the actual position after all operations

    # Calculate dealer section position:
    # - Original template: dealer section at row 47
    # - After inserting rows_to_insert: row 47 + rows_to_insert
    # - After deleting 21 rows (rows 24-44): row 47 + rows_to_insert - 21 = 26 + rows_to_insert
    DEALER_SECTION_START_ROW = 26 + rows_to_insert
    DEALER_HEADER_ROW = 27 + rows_to_insert
    DEALER_DATA_START_ROW = 28 + rows_to_insert
    DEALER_DATA_END_ROW = 65 + rows_to_insert

    # Collect all dealers from sessions
    dealers_with_rake: dict[int, list[DealerRakeEntry]] = {}
    dealer_names: dict[int, str] = {}

    for session in sessions:
        for assignment in session.dealer_assignments:
            dealer_id = int(cast(int, assignment.dealer_id))
            if assignment.dealer:
                dealer_names[dealer_id] = cast(str, assignment.dealer.username)
            if dealer_id not in dealers_with_rake:
                dealers_with_rake[dealer_id] = []
            for entry in (assignment.rake_entries or []):
                dealers_with_rake[dealer_id].append(entry)

    # Sort dealers by ID
    sorted_dealer_ids = sorted(dealers_with_rake.keys())
    num_dealers = len(sorted_dealer_ids)

    # Dealer section now starts at column K (11)
    # Dealers use columns K-L, M-N, O-P, ... (11-12, 13-14, 15-16, ...)
    no_border = Border()
    TEMPLATE_DEALER_CLEAR_END = 70 + rows_to_insert  # Clear well past the template summary rows

    # Clear old template data in columns A-Z (1-26) in the bottom section area
    # Template has dealer data in A-F and formulas in K-P that need to be cleared
    for row in range(DEALER_SECTION_START_ROW, TEMPLATE_DEALER_CLEAR_END + 1):
        for col in range(1, 27):  # A through Z (1-26)
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.fill = no_fill
            cell.border = no_border

    # Define dealer block background colors
    dealer_name_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Blue for name
    dealer_header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")  # Light blue for headers
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Track the maximum data row used across all dealers
    max_dealer_data_row = DEALER_DATA_START_ROW - 1  # Start before first data row

    # Fill dealer names, headers, and rake entries (only for actual dealers)
    for idx, dealer_id in enumerate(sorted_dealer_ids):
        col_rake = 11 + idx * 2   # K=11, M=13, O=15, ...
        col_time = col_rake + 1   # L=12, N=14, P=16, ...

        # Dealer name with background color
        dealer_name = dealer_names.get(dealer_id, f"D{idx+1}")
        name_cell = ws.cell(row=DEALER_SECTION_START_ROW, column=col_rake, value=dealer_name)
        name_cell.fill = dealer_name_fill
        name_cell.font = Font(bold=True, color="FFFFFF")
        name_cell.border = thin_border
        # Extend background to second column of dealer block
        name_cell2 = ws.cell(row=DEALER_SECTION_START_ROW, column=col_time)
        name_cell2.fill = dealer_name_fill
        name_cell2.border = thin_border

        # Headers with background color
        header_rake = ws.cell(row=DEALER_HEADER_ROW, column=col_rake, value="Рейк")
        header_rake.fill = dealer_header_fill
        header_rake.border = thin_border
        header_time = ws.cell(row=DEALER_HEADER_ROW, column=col_time, value="Время")
        header_time.fill = dealer_header_fill
        header_time.border = thin_border

        # Rake entries
        rake_entries = sorted(dealers_with_rake[dealer_id], key=lambda e: cast(dt.datetime, e.created_at))
        for i, entry in enumerate(rake_entries):
            row = DEALER_DATA_START_ROW + i
            if row > DEALER_DATA_END_ROW:
                break
            rake_cell = ws.cell(row=row, column=col_rake, value=int(cast(int, entry.amount)))
            rake_cell.border = thin_border
            time_str = cast(dt.datetime, entry.created_at).strftime("%H:%M")
            time_cell = ws.cell(row=row, column=col_time, value=time_str)
            time_cell.border = thin_border
            # Track the maximum row used
            if row > max_dealer_data_row:
                max_dealer_data_row = row

    # === DEALER TOTALS SECTION ===
    # Position totals right after the last data row (not at a fixed position)
    # Add 1 empty row after data, then totals
    grand_total_rake = 0

    # Border with thick top for totals row
    total_top_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='medium'),
        bottom=Side(style='thin')
    )

    if num_dealers > 0:
        DEALER_TOTALS_ROW = max_dealer_data_row + 2  # 1 empty row, then totals

        # Add per-dealer totals (only for dealers that have data)
        for idx, dealer_id in enumerate(sorted_dealer_ids):
            col_rake = 11 + idx * 2   # K=11, M=13, O=15, ...
            col_time = col_rake + 1   # L=12, N=14, P=16, ...
            col_letter = get_column_letter(col_rake)

            # Calculate total for this dealer
            rake_entries = dealers_with_rake[dealer_id]
            dealer_total = sum(int(cast(int, entry.amount)) for entry in rake_entries)
            grand_total_rake += dealer_total

            # Write SUM formula for this dealer with top border
            total_cell = ws.cell(row=DEALER_TOTALS_ROW, column=col_rake,
                    value=f"=SUM({col_letter}{DEALER_DATA_START_ROW}:{col_letter}{max_dealer_data_row})")
            total_cell.font = Font(bold=True)
            total_cell.border = total_top_border

            # Add top border to time column too (spanning both columns)
            time_total_cell = ws.cell(row=DEALER_TOTALS_ROW, column=col_time)
            time_total_cell.border = total_top_border

    # === SECTION: Clear columns beyond dealer area for bottom part ===
    # Dealers are in columns K+ (11+), so we need to clear columns well beyond
    # where dealers could be. With up to 10 dealers, that's columns 11-30.
    # Clear columns 31+ (AE onwards) for safety
    no_border = Border()
    BOTTOM_SECTION_START = 26 + rows_to_insert
    BOTTOM_SECTION_END = BOTTOM_SECTION_START + 50  # Clear enough rows for bottom section

    for row in range(BOTTOM_SECTION_START, BOTTOM_SECTION_END):
        for col in range(31, 50):  # AE onwards (column 31+) - well after dealer columns
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.fill = no_fill
            cell.border = no_border

    # === Clear W-Z columns in seat area (columns 23-26, rows 1 to summary) ===
    for row in range(1, summary_start_row + 7):
        for col in range(23, 27):  # W, X, Y, Z
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.fill = no_fill
            cell.border = no_border

    # === Clear area below seats but before bottom blocks ===
    # This is the gap between summary rows and dealer section
    # Don't clear BOTTOM_SECTION_START itself - that's where dealer names go
    gap_start = summary_start_row + 6
    gap_end = BOTTOM_SECTION_START - 1  # Stop before dealer name row
    for row in range(gap_start, gap_end + 1):
        for col in range(1, 50):  # Clear all columns in the gap area
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.fill = no_fill
            cell.border = no_border

    # === SECTION: B-D columns - Расходы, Доходы, З/П тотал, Рейк ===
    # Data starts at column B (2) - fixed position for these blocks
    # Use template font: Roboto Mono, size 11

    # After operations, template row 48 becomes 27 + rows_to_insert
    NP_SECTION_START = 27 + rows_to_insert
    template_font = Font(name="Roboto Mono", size=11)
    template_font_bold = Font(name="Roboto Mono", size=11, bold=True)

    # Clear old template data in B-J columns (2-10) for bottom section
    # Columns K+ (11+) are used by dealer section which is already written above
    # Start from row 26 (dealer name row) to clear any leftover data
    no_border = Border()
    TEMPLATE_CLEAR_START = 26 + rows_to_insert
    TEMPLATE_CLEAR_END = 70 + rows_to_insert

    for clear_row in range(TEMPLATE_CLEAR_START, TEMPLATE_CLEAR_END + 1):
        for col in range(2, 11):  # B through J columns (2-10), skip K+ where dealers are
            cell = ws.cell(row=clear_row, column=col)
            cell.value = None
            cell.fill = no_fill
            cell.border = no_border
            cell.font = Font()
            cell.alignment = Alignment()

    # Prepare data
    negative_adjustments = [ba for ba in balance_adjustments if int(cast(int, ba.amount)) < 0]
    positive_adjustments = [ba for ba in balance_adjustments if int(cast(int, ba.amount)) > 0]

    # Calculate staff salaries
    staff_salaries: list[tuple[str, str, int]] = []  # (name, role, earnings)
    total_staff_salary = 0

    dealers = [s for s in staff if s.role == "dealer"]
    for dealer in dealers:
        dealer_id = int(cast(int, dealer.id))
        dealer_name = cast(str, dealer.username)
        hourly_rate = int(cast(int, dealer.hourly_rate)) if dealer.hourly_rate else 0
        hours = _calculate_dealer_hours(sessions, dealer_id)
        if hours > 0:
            earnings = round(hours * hourly_rate)
            total_staff_salary += earnings
            staff_salaries.append((dealer_name, "Дилер", earnings))

    waiters = [s for s in staff if s.role == "waiter"]
    for waiter in waiters:
        waiter_id = int(cast(int, waiter.id))
        waiter_name = cast(str, waiter.username)
        hourly_rate = int(cast(int, waiter.hourly_rate)) if waiter.hourly_rate else 0
        hours = _calculate_waiter_hours(sessions, waiter_id)
        if hours > 0:
            earnings = round(hours * hourly_rate)
            total_staff_salary += earnings
            staff_salaries.append((waiter_name, "Официант", earnings))

    # Write all data starting from NP_SECTION_START, column H (8)
    current_row = NP_SECTION_START
    stats_start_row = current_row  # Track start for outer border

    # Define light background colors for each section
    expenses_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Light yellow
    income_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Light green
    salary_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # Light blue
    rake_fill = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")  # Light gray
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # === РАСХОДЫ (Expenses) ===
    expenses_start = current_row
    ws.cell(row=current_row, column=2, value="РАСХОДЫ")
    ws.cell(row=current_row, column=2).font = template_font_bold
    current_row += 1

    if negative_adjustments:
        for adj in negative_adjustments:
            amount = int(cast(int, adj.amount))
            comment = cast(str, adj.comment) if adj.comment else ""
            ws.cell(row=current_row, column=2, value=comment[:30])
            ws.cell(row=current_row, column=2).font = template_font
            ws.cell(row=current_row, column=3, value=amount)
            ws.cell(row=current_row, column=3).font = template_font
            current_row += 1
        expenses_total = sum(int(cast(int, ba.amount)) for ba in negative_adjustments)
        ws.cell(row=current_row, column=2, value="Итого:")
        ws.cell(row=current_row, column=2).font = template_font_bold
        ws.cell(row=current_row, column=3, value=expenses_total)
        ws.cell(row=current_row, column=3).font = template_font_bold
        current_row += 1
    else:
        ws.cell(row=current_row, column=2, value="(нет)")
        ws.cell(row=current_row, column=2).font = template_font
        current_row += 1
    expenses_end = current_row - 1

    # Apply expenses background
    for r in range(expenses_start, expenses_end + 1):
        for c in range(2, 5):  # B-D
            ws.cell(row=r, column=c).fill = expenses_fill
            ws.cell(row=r, column=c).border = thin_border

    current_row += 1  # Empty row

    # === ДОХОДЫ (Income) ===
    income_start = current_row
    ws.cell(row=current_row, column=2, value="ДОХОДЫ")
    ws.cell(row=current_row, column=2).font = template_font_bold
    current_row += 1

    if positive_adjustments:
        for adj in positive_adjustments:
            amount = int(cast(int, adj.amount))
            comment = cast(str, adj.comment) if adj.comment else ""
            ws.cell(row=current_row, column=2, value=comment[:30])
            ws.cell(row=current_row, column=2).font = template_font
            ws.cell(row=current_row, column=3, value=amount)
            ws.cell(row=current_row, column=3).font = template_font
            current_row += 1
        bonuses_total = sum(int(cast(int, ba.amount)) for ba in positive_adjustments)
        ws.cell(row=current_row, column=2, value="Итого:")
        ws.cell(row=current_row, column=2).font = template_font_bold
        ws.cell(row=current_row, column=3, value=bonuses_total)
        ws.cell(row=current_row, column=3).font = template_font_bold
        current_row += 1
    else:
        ws.cell(row=current_row, column=2, value="(нет)")
        ws.cell(row=current_row, column=2).font = template_font
        current_row += 1
    income_end = current_row - 1

    # Apply income background
    for r in range(income_start, income_end + 1):
        for c in range(2, 5):  # B-D
            ws.cell(row=r, column=c).fill = income_fill
            ws.cell(row=r, column=c).border = thin_border

    current_row += 1  # Empty row

    # === З/П ТОТАЛ (Staff Salaries) ===
    salary_start = current_row
    ws.cell(row=current_row, column=2, value="З/П ТОТАЛ")
    ws.cell(row=current_row, column=2).font = template_font_bold
    current_row += 1

    if staff_salaries:
        for name, role, earnings in staff_salaries:
            ws.cell(row=current_row, column=2, value=name)
            ws.cell(row=current_row, column=2).font = template_font
            ws.cell(row=current_row, column=3, value=role)
            ws.cell(row=current_row, column=3).font = template_font
            ws.cell(row=current_row, column=4, value=earnings)
            ws.cell(row=current_row, column=4).font = template_font
            current_row += 1
        ws.cell(row=current_row, column=2, value="Итого:")
        ws.cell(row=current_row, column=2).font = template_font_bold
        ws.cell(row=current_row, column=4, value=total_staff_salary)
        ws.cell(row=current_row, column=4).font = template_font_bold
        current_row += 1
    else:
        ws.cell(row=current_row, column=2, value="(нет)")
        ws.cell(row=current_row, column=2).font = template_font
        current_row += 1
    salary_end = current_row - 1

    # Apply salary background
    for r in range(salary_start, salary_end + 1):
        for c in range(2, 5):  # B-D
            ws.cell(row=r, column=c).fill = salary_fill
            ws.cell(row=r, column=c).border = thin_border

    current_row += 1  # Blank line

    # === РЕЙК БРУТТО / РЕЙК НЕТТО ===
    rake_start = current_row
    ws.cell(row=current_row, column=2, value="Рейк брутто")
    ws.cell(row=current_row, column=2).font = template_font_bold
    ws.cell(row=current_row, column=3, value=grand_total_rake)
    ws.cell(row=current_row, column=3).font = template_font_bold
    current_row += 1

    # Calculate totals for net rake
    # total_expenses: sum of negative balance adjustments (already negative values)
    total_expenses = sum(int(cast(int, ba.amount)) for ba in negative_adjustments) if negative_adjustments else 0
    # total_income: sum of positive balance adjustments
    total_income = sum(int(cast(int, ba.amount)) for ba in positive_adjustments) if positive_adjustments else 0

    # Net rake = rake brutto - expenses + income - salaries
    # Since total_expenses is already negative, we add it (which subtracts the expense)
    net_rake = grand_total_rake + total_expenses + total_income - total_staff_salary

    ws.cell(row=current_row, column=2, value="Рейк нетто")
    ws.cell(row=current_row, column=2).font = template_font_bold
    ws.cell(row=current_row, column=3, value=net_rake)
    ws.cell(row=current_row, column=3).font = template_font_bold
    current_row += 1
    rake_end = current_row - 1

    # Apply rake background
    for r in range(rake_start, rake_end + 1):
        for c in range(2, 5):  # B-D
            ws.cell(row=r, column=c).fill = rake_fill
            ws.cell(row=r, column=c).border = thin_border

    stats_end_row = current_row - 1  # Track end for outer border

    # Apply thick outer border around entire stats section
    thick_side = Side(style='medium')
    for r in range(stats_start_row, stats_end_row + 1):
        for c in range(2, 5):  # B-D
            cell = ws.cell(row=r, column=c)
            left = thick_side if c == 2 else cell.border.left
            right = thick_side if c == 4 else cell.border.right
            top = thick_side if r == stats_start_row else cell.border.top
            bottom = thick_side if r == stats_end_row else cell.border.bottom
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    # === SECTION: F-I columns - Chip operations (+/-) ===
    # Data starts at column F (6) - Жетоны section
    # F = (-) amount, G = (-) time, H = (+) amount, I = (+) time

    QT_SECTION_START = 27 + rows_to_insert  # Template row 48

    # Note: B-U columns already cleared above

    # Define background colors for chip operations
    negative_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # Very light red/peach
    positive_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Light green
    chip_thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Header row - columns F-I (6-9) with title spanning across
    chips_start_row = QT_SECTION_START
    ws.cell(row=QT_SECTION_START, column=7, value="Жетоны")
    ws.cell(row=QT_SECTION_START, column=7).font = template_font_bold
    ws.cell(row=QT_SECTION_START, column=8, value="на столе")
    ws.cell(row=QT_SECTION_START, column=8).font = template_font_bold
    # Apply borders to header row
    for c in range(6, 10):
        ws.cell(row=QT_SECTION_START, column=c).border = chip_thin_border

    # Column headers row (2 rows down)
    headers_row = QT_SECTION_START + 2
    ws.cell(row=headers_row, column=6, value="(-)")
    ws.cell(row=headers_row, column=6).font = template_font_bold
    ws.cell(row=headers_row, column=6).fill = negative_fill
    ws.cell(row=headers_row, column=7).fill = negative_fill
    ws.cell(row=headers_row, column=8, value="(+)")
    ws.cell(row=headers_row, column=8).font = template_font_bold
    ws.cell(row=headers_row, column=8).fill = positive_fill
    ws.cell(row=headers_row, column=9).fill = positive_fill
    # Apply borders to headers row
    for c in range(6, 10):
        ws.cell(row=headers_row, column=c).border = chip_thin_border

    # Separate chip purchases into negative (cashouts) and positive (buy-ins)
    negative_ops: list[tuple[int, dt.datetime]] = []  # (amount, timestamp)
    positive_ops: list[tuple[int, dt.datetime]] = []  # (amount, timestamp)

    for p in purchases:
        amount = int(cast(int, p.amount))
        ts = cast(dt.datetime, p.created_at)
        if amount < 0:
            negative_ops.append((amount, ts))
        elif amount > 0:
            positive_ops.append((amount, ts))

    # Sort by timestamp
    negative_ops.sort(key=lambda x: x[1])
    positive_ops.sort(key=lambda x: x[1])

    # Write data rows
    data_start_row = headers_row + 1
    max_ops = max(len(negative_ops), len(positive_ops)) if negative_ops or positive_ops else 0

    for i in range(max_ops):
        row = data_start_row + i

        # Negative (cashout) - columns F (6) and G (7)
        cell_F = ws.cell(row=row, column=6)
        cell_G = ws.cell(row=row, column=7)
        cell_F.fill = negative_fill
        cell_F.border = chip_thin_border
        cell_G.fill = negative_fill
        cell_G.border = chip_thin_border
        if i < len(negative_ops):
            amount, ts = negative_ops[i]
            cell_F.value = amount
            cell_F.font = template_font
            cell_G.value = ts.strftime("%H:%M")
            cell_G.font = template_font

        # Positive (buy-in) - columns H (8) and I (9)
        cell_H = ws.cell(row=row, column=8)
        cell_I = ws.cell(row=row, column=9)
        cell_H.fill = positive_fill
        cell_H.border = chip_thin_border
        cell_I.fill = positive_fill
        cell_I.border = chip_thin_border
        if i < len(positive_ops):
            amount, ts = positive_ops[i]
            cell_H.value = amount
            cell_H.font = template_font
            cell_I.value = ts.strftime("%H:%M")
            cell_I.font = template_font

    # Totals row
    chips_end_row = data_start_row + max_ops - 1 if max_ops > 0 else headers_row
    if max_ops > 0:
        totals_row = data_start_row + max_ops
        neg_total = sum(op[0] for op in negative_ops)
        pos_total = sum(op[0] for op in positive_ops)

        # Negative total with label
        cell_F_tot = ws.cell(row=totals_row, column=6, value=neg_total)
        cell_F_tot.font = template_font_bold
        cell_F_tot.fill = negative_fill
        cell_F_tot.border = chip_thin_border
        cell_G_tot = ws.cell(row=totals_row, column=7, value="Σ(-)")
        cell_G_tot.font = template_font_bold
        cell_G_tot.fill = negative_fill
        cell_G_tot.border = chip_thin_border

        # Positive total with label
        cell_H_tot = ws.cell(row=totals_row, column=8, value=pos_total)
        cell_H_tot.font = template_font_bold
        cell_H_tot.fill = positive_fill
        cell_H_tot.border = chip_thin_border
        cell_I_tot = ws.cell(row=totals_row, column=9, value="Σ(+)")
        cell_I_tot.font = template_font_bold
        cell_I_tot.fill = positive_fill
        cell_I_tot.border = chip_thin_border

        # Net change row with label
        net_row = totals_row + 1
        ws.cell(row=net_row, column=6).border = chip_thin_border
        cell_net_label = ws.cell(row=net_row, column=7, value="Итого:")
        cell_net_label.font = template_font_bold
        cell_net_label.border = chip_thin_border
        cell_net = ws.cell(row=net_row, column=8, value=neg_total + pos_total)
        cell_net.font = template_font_bold
        cell_net.border = chip_thin_border
        ws.cell(row=net_row, column=9).border = chip_thin_border
        chips_end_row = net_row

    # Apply outer border to entire chip operations section
    thick_side = Side(style='medium')
    for r in range(chips_start_row, chips_end_row + 1):
        for c in range(6, 10):  # F-I
            cell = ws.cell(row=r, column=c)
            left = thick_side if c == 6 else cell.border.left
            right = thick_side if c == 9 else cell.border.right
            top = thick_side if r == chips_start_row else cell.border.top
            bottom = thick_side if r == chips_end_row else cell.border.bottom
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)


def _get_player_blocks_for_session(session_id: str, db: DBSession) -> list[dict]:
    """
    Get player blocks for a session.
    A player block is a unique combination of (seat_no, player_name).
    If a player leaves and re-enters the same seat, their actions are compiled into one block.
    If a player moves to a different seat, that's a separate block.

    Returns list of dicts with:
    - seat_no: int
    - player_name: str
    - chip_history: list of dicts with 'amount' and 'timestamp' for each chip operation
    - chips_taken: int (sum of all positive chip purchases)
    - chips_returned: int (absolute value of sum of all negative chip operations - cashouts)
    - result: int (chips_taken - chips_returned, can be positive or negative)
    - cash_given: int (sum of all positive cash purchases)
    """
    # Get all name changes for this session to track player history
    name_changes = (
        db.query(SeatNameChange)
        .filter(SeatNameChange.session_id == session_id)
        .order_by(SeatNameChange.created_at.asc())
        .all()
    )

    # Get all chip purchases for this session
    purchases = (
        db.query(ChipPurchase)
        .filter(ChipPurchase.session_id == session_id)
        .order_by(ChipPurchase.created_at.asc())
        .all()
    )

    # Get all chip operations (includes cashouts which don't have purchases)
    chip_ops = (
        db.query(ChipOp)
        .filter(ChipOp.session_id == session_id)
        .order_by(ChipOp.created_at.asc())
        .all()
    )

    # Build a map of chip_op_id to purchase for quick lookup
    purchase_by_op_id = {int(cast(int, p.chip_op_id)): p for p in purchases}

    # Track player blocks: key is (seat_no, player_name), value is the block data
    player_blocks: dict[tuple[int, str], dict] = {}

    # Build a timeline of events (name changes and chip ops) to properly track player at each seat
    events = []

    # Add name changes to timeline
    for nc in name_changes:
        events.append({
            "type": "name_change",
            "timestamp": cast(dt.datetime, nc.created_at),
            "seat_no": int(cast(int, nc.seat_no)),
            "old_name": nc.old_name,
            "new_name": nc.new_name,
            "change_type": cast(str, nc.change_type) if nc.change_type else "name_change",
        })

    # Add chip operations to timeline
    for op in chip_ops:
        events.append({
            "type": "chip_op",
            "timestamp": cast(dt.datetime, op.created_at),
            "seat_no": int(cast(int, op.seat_no)),
            "amount": int(cast(int, op.amount)),
            "op_id": int(cast(int, op.id)),
        })

    # Sort events by timestamp
    events.sort(key=lambda x: x["timestamp"])

    # Track current player at each seat
    current_player_at_seat: dict[int, str | None] = {}

    # Process events in chronological order
    for event in events:
        seat_no = event["seat_no"]

        if event["type"] == "name_change":
            change_type = event["change_type"]
            new_name = event["new_name"]

            if change_type == "player_left":
                # Player left, clear the seat
                current_player_at_seat[seat_no] = None
            else:
                # Name change or new player
                current_player_at_seat[seat_no] = new_name

                # Initialize player block if this is a new (seat_no, player_name) combination
                if new_name:
                    block_key = (seat_no, new_name)
                    if block_key not in player_blocks:
                        player_blocks[block_key] = {
                            "seat_no": seat_no,
                            "player_name": new_name,
                            "chip_history": [],
                            "chips_taken": 0,
                            "chips_returned": 0,
                            "result": 0,
                            "cash_given": 0,
                        }

        elif event["type"] == "chip_op":
            amount = event["amount"]
            op_id = event["op_id"]

            # Get the player who is currently at this seat
            player_name = current_player_at_seat.get(seat_no)

            if player_name:
                block_key = (seat_no, player_name)

                # Ensure block exists (it should have been created by a name_change event)
                if block_key not in player_blocks:
                    player_blocks[block_key] = {
                        "seat_no": seat_no,
                        "player_name": player_name,
                        "chip_history": [],
                        "chips_taken": 0,
                        "chips_returned": 0,
                        "result": 0,
                        "cash_given": 0,
                    }

                block = player_blocks[block_key]

                # Add to chip history
                block["chip_history"].append({
                    "amount": amount,
                    "timestamp": event["timestamp"],
                })

                # Get purchase info if it exists
                purchase = purchase_by_op_id.get(op_id)

                if amount > 0:
                    # Positive amount = chips taken (buyin)
                    block["chips_taken"] += amount

                    # If it's a cash purchase, add to cash_given
                    if purchase and cast(str, purchase.payment_type) == "cash":
                        block["cash_given"] += amount
                elif amount < 0:
                    # Negative amount = chips returned (cashout)
                    block["chips_returned"] += abs(amount)

    # Calculate result for each block
    for block in player_blocks.values():
        block["result"] = block["chips_taken"] - block["chips_returned"]

    # Convert to list and sort by seat_no, then player_name
    blocks_list = sorted(player_blocks.values(), key=lambda x: (x["seat_no"], x["player_name"]))

    return blocks_list


def _create_table_states_sheet(
    wb: Workbook,
    tables: list[Table],
    sessions: list[Session],
    seats_by_session: dict[str, list[Seat]],
    db,
):
    """Create sheet with table states - seats, players, totals, and staff earnings."""
    ws = wb.create_sheet(title="Состояние столов")

    if not sessions:
        ws.cell(row=1, column=1, value="Нет данных за выбранную дату")
        ws.cell(row=1, column=1).font = Font(italic=True)
        return

    # Define visual styles for tables and sessions
    table_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Blue
    session_header_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")  # Light gray
    thick_border = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    )
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    row = 1
    for table in tables:
        table_sessions = [s for s in sessions if s.table_id == table.id]
        if not table_sessions:
            continue

        # Table header with blue background and thick border
        table_header_cell = ws.cell(row=row, column=1, value=f"Стол: {table.name}")
        table_header_cell.font = Font(bold=True, size=14, color="FFFFFF")
        table_header_cell.fill = table_header_fill
        table_header_cell.border = thick_border
        # Merge cells for table header to span across columns (7 columns now)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = table_header_fill
            ws.cell(row=row, column=col).border = thick_border
        row += 1

        for session in table_sessions:
            session_start_row = row  # Track where this session starts
            sid = cast(str, session.id)
            seats = seats_by_session.get(sid, [])

            # Session info with gray background and thin border
            start_time = cast(dt.datetime, session.created_at).strftime("%H:%M")
            if session.closed_at:
                end_time = cast(dt.datetime, session.closed_at).strftime("%H:%M")
            elif session.status == "closed":
                end_time = "закрыта"
            else:
                end_time = "открыта"

            # Show all dealers if there are multiple concurrent dealers
            if session.dealer_assignments and len(session.dealer_assignments) > 0:
                active_dealers = [a for a in session.dealer_assignments if not a.ended_at]
                if len(active_dealers) > 1:
                    dealer_names = ", ".join([a.dealer.username if a.dealer else "—" for a in active_dealers])
                    dealer_name = f"{dealer_names} (несколько)"
                elif len(active_dealers) == 1:
                    dealer_name = active_dealers[0].dealer.username if active_dealers[0].dealer else "—"
                else:
                    # All dealers ended, show the last one
                    dealer_name = session.dealer.username if session.dealer else "—"
            else:
                dealer_name = session.dealer.username if session.dealer else "—"

            waiter_name = session.waiter.username if session.waiter else "—"
            status_text = "закрыта" if session.status == "closed" else "открыта"
            chips_in_play = int(cast(int, session.chips_in_play))

            # Session header row with light gray background
            session_cells = [
                (1, f"Сессия: {start_time} - {end_time}"),
                (2, f"Дилер: {dealer_name}"),
                (3, f"Официант: {waiter_name}"),
                (4, f"Статус: {status_text}")
            ]
            for col, value in session_cells:
                cell = ws.cell(row=row, column=col, value=value)
                cell.fill = session_header_fill
                cell.border = thin_border
                cell.font = Font(bold=True)
            row += 1

            # Chips in play info
            ws.cell(row=row, column=1, value=f"Фишек на столе: {chips_in_play}")
            row += 1

            # Player blocks header - 7 columns total
            headers = ["Место", "Игрок", "История", "Жетоны(-)", "Жетоны(+)", "Результат", "Наличных"]
            for col, h in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=h)
            _style_header(ws, row, len(headers))
            row += 1

            # Get player blocks for this session
            player_blocks = _get_player_blocks_for_session(sid, db)

            # Display player blocks
            session_total_result = 0
            for block in player_blocks:
                # Only show blocks with activity
                if block["chips_taken"] > 0 or block["chips_returned"] > 0:
                    # Determine how many rows this block needs (one per history entry)
                    history = block.get("chip_history", [])
                    num_history_rows = max(1, len(history))

                    block_start_row = row

                    # Column 1: Seat number (merged across all history rows)
                    cell_seat = ws.cell(row=row, column=1, value=block["seat_no"])
                    if num_history_rows > 1:
                        ws.merge_cells(start_row=row, start_column=1, end_row=row + num_history_rows - 1, end_column=1)
                        # Apply to all merged cells for consistency
                        for i in range(num_history_rows):
                            ws.cell(row=row + i, column=1).border = thin_border

                    # Column 2: Player name (merged across all history rows)
                    cell_player = ws.cell(row=row, column=2, value=block["player_name"])
                    if num_history_rows > 1:
                        ws.merge_cells(start_row=row, start_column=2, end_row=row + num_history_rows - 1, end_column=2)
                        for i in range(num_history_rows):
                            ws.cell(row=row + i, column=2).border = thin_border

                    # Column 3: History - one row per chip operation
                    for i, hist_entry in enumerate(history):
                        hist_row = row + i
                        amount = hist_entry["amount"]
                        timestamp = hist_entry["timestamp"].strftime("%H:%M")
                        cell_hist = ws.cell(row=hist_row, column=3, value=f"{amount:+d} {timestamp}")
                        cell_hist.border = thin_border

                    # Column 4: Chips taken (Orange) - merged
                    cell_taken = ws.cell(row=row, column=4, value=block["chips_taken"])
                    cell_taken.fill = CHIPS_TAKEN_FILL
                    cell_taken.border = thin_border
                    if num_history_rows > 1:
                        ws.merge_cells(start_row=row, start_column=4, end_row=row + num_history_rows - 1, end_column=4)
                        # Apply fill and border to all merged cells
                        for i in range(num_history_rows):
                            ws.cell(row=row + i, column=4).fill = CHIPS_TAKEN_FILL
                            ws.cell(row=row + i, column=4).border = thin_border

                    # Column 5: Chips returned (Cyan) - merged
                    cell_returned = ws.cell(row=row, column=5, value=block["chips_returned"])
                    cell_returned.fill = CHIPS_RETURNED_FILL
                    cell_returned.border = thin_border
                    if num_history_rows > 1:
                        ws.merge_cells(start_row=row, start_column=5, end_row=row + num_history_rows - 1, end_column=5)
                        for i in range(num_history_rows):
                            ws.cell(row=row + i, column=5).fill = CHIPS_RETURNED_FILL
                            ws.cell(row=row + i, column=5).border = thin_border

                    # Column 6: Result (Magenta) - merged
                    cell_result = ws.cell(row=row, column=6, value=block["result"])
                    cell_result.fill = RESULT_FILL
                    cell_result.border = thin_border
                    if num_history_rows > 1:
                        ws.merge_cells(start_row=row, start_column=6, end_row=row + num_history_rows - 1, end_column=6)
                        for i in range(num_history_rows):
                            ws.cell(row=row + i, column=6).fill = RESULT_FILL
                            ws.cell(row=row + i, column=6).border = thin_border

                    # Column 7: Cash given (Yellow) - merged
                    cell_cash = ws.cell(row=row, column=7, value=block["cash_given"])
                    cell_cash.fill = CASH_GIVEN_FILL
                    cell_cash.border = thin_border
                    if num_history_rows > 1:
                        ws.merge_cells(start_row=row, start_column=7, end_row=row + num_history_rows - 1, end_column=7)
                        for i in range(num_history_rows):
                            ws.cell(row=row + i, column=7).fill = CASH_GIVEN_FILL
                            ws.cell(row=row + i, column=7).border = thin_border

                    session_total_result += block["result"]
                    row += num_history_rows

            # Session total
            ws.cell(row=row, column=5, value="ИТОГО сессии:")
            ws.cell(row=row, column=5).font = Font(bold=True)
            cell = ws.cell(row=row, column=6, value=session_total_result)
            cell.font = Font(bold=True)
            if session_total_result > 0:
                cell.fill = MONEY_POSITIVE_FILL
            elif session_total_result < 0:
                cell.fill = MONEY_NEGATIVE_FILL
            row += 1

            # Staff earnings section with rake
            dealer_earnings = _calculate_session_dealer_earnings(session, db)
            waiter_earnings = _calculate_session_waiter_earnings(session, db)

            if dealer_earnings or waiter_earnings:
                row += 1  # Add spacing
                ws.cell(row=row, column=1, value="Персонал сессии:")
                ws.cell(row=row, column=1).font = Font(bold=True, italic=True)
                row += 1

                # Dealer earnings header (now with Rake column)
                if dealer_earnings:
                    staff_headers = ["Сотрудник", "Роль", "Часов", "Ставка/час", "Зарплата", "Рейк"]
                    for col, h in enumerate(staff_headers, 1):
                        ws.cell(row=row, column=col, value=h)
                    _style_header(ws, row, len(staff_headers))
                    row += 1

                    # Display each dealer's earnings with rake
                    total_dealer_salary = 0
                    total_dealer_rake = 0
                    for earning in dealer_earnings:
                        ws.cell(row=row, column=1, value=earning["dealer_name"])
                        ws.cell(row=row, column=2, value="Дилер")
                        ws.cell(row=row, column=3, value=round(earning["hours"], 2))
                        ws.cell(row=row, column=4, value=earning["hourly_rate"])
                        salary_cell = ws.cell(row=row, column=5, value=earning["salary"])
                        salary_cell.fill = MONEY_NEGATIVE_FILL  # Salary is an expense
                        total_dealer_salary += earning["salary"]
                        # Rake column
                        rake_value = earning.get("rake", 0) or 0
                        rake_cell = ws.cell(row=row, column=6, value=rake_value)
                        if rake_value > 0:
                            rake_cell.fill = MONEY_POSITIVE_FILL
                        total_dealer_rake += rake_value
                        row += 1

                    # Show total if multiple dealers
                    if len(dealer_earnings) > 1:
                        ws.cell(row=row, column=4, value="Итого дилеры:")
                        ws.cell(row=row, column=4).font = Font(bold=True)
                        total_cell = ws.cell(row=row, column=5, value=total_dealer_salary)
                        total_cell.font = Font(bold=True)
                        total_cell.fill = MONEY_NEGATIVE_FILL
                        total_rake_cell = ws.cell(row=row, column=6, value=total_dealer_rake)
                        total_rake_cell.font = Font(bold=True)
                        if total_dealer_rake > 0:
                            total_rake_cell.fill = MONEY_POSITIVE_FILL
                        row += 1

                # Waiter earnings
                if waiter_earnings:
                    # Add header if not already added
                    if not dealer_earnings:
                        staff_headers = ["Сотрудник", "Роль", "Часов", "Ставка/час", "Зарплата", "Рейк"]
                        for col, h in enumerate(staff_headers, 1):
                            ws.cell(row=row, column=col, value=h)
                        _style_header(ws, row, len(staff_headers))
                        row += 1

                    ws.cell(row=row, column=1, value=waiter_earnings["waiter_name"])
                    ws.cell(row=row, column=2, value="Официант")
                    ws.cell(row=row, column=3, value=round(waiter_earnings["hours"], 2))
                    ws.cell(row=row, column=4, value=waiter_earnings["hourly_rate"])
                    salary_cell = ws.cell(row=row, column=5, value=waiter_earnings["salary"])
                    salary_cell.fill = MONEY_NEGATIVE_FILL  # Salary is an expense
                    ws.cell(row=row, column=6, value="—")  # Waiters don't have rake
                    row += 1

            # Apply border around the entire session block
            session_end_row = row - 1
            _apply_session_border(ws, session_start_row, session_end_row, max_col=7)

            row += 1  # Extra space after session

        row += 1  # Extra space between tables

    _auto_width(ws)


def _create_purchases_sheet(
    wb: Workbook,
    purchases: list[ChipPurchase],
    tables: list[Table],
    db: DBSession,
):
    """Create sheet with chip purchase chronology."""
    ws = wb.create_sheet(title="Хронология покупок")

    # Headers
    headers = ["Время", "Стол", "Место", "Сумма", "Тип оплаты", "Выдал"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, len(headers))

    if not purchases:
        ws.cell(row=2, column=1, value="Нет покупок за выбранную дату")
        ws.cell(row=2, column=1).font = Font(italic=True)
        _auto_width(ws)
        return

    tables_by_id = {t.id: t for t in tables}

    row = 2
    for p in purchases:
        time_str = cast(dt.datetime, p.created_at).strftime("%H:%M:%S")
        table = tables_by_id.get(int(cast(int, p.table_id)))
        table_name = cast(str, table.name) if table else f"ID {p.table_id}"

        ws.cell(row=row, column=1, value=time_str)
        ws.cell(row=row, column=2, value=table_name)
        ws.cell(row=row, column=3, value=int(cast(int, p.seat_no)))

        amount = int(cast(int, p.amount))
        cell = ws.cell(row=row, column=4, value=amount)
        # For cashouts (negative), show as expense (red)
        # For buyins (positive), show as income (green)
        if amount > 0:
            cell.fill = MONEY_POSITIVE_FILL
        elif amount < 0:
            cell.fill = MONEY_NEGATIVE_FILL

        # Payment type column
        # For cashouts, show "выдача" (payout) instead of payment type
        if amount < 0:
            payment_text = "выдача"
            payment_cell = ws.cell(row=row, column=5, value=payment_text)
            payment_cell.fill = MONEY_NEGATIVE_FILL
            payment_cell.font = Font(bold=True)
        else:
            payment_type = cast(str, p.payment_type) if p.payment_type else "cash"
            payment_text = "наличные" if payment_type == "cash" else "кредит"
            payment_cell = ws.cell(row=row, column=5, value=payment_text)
            # Apply dark color coding for payment type
            if payment_type == "cash":
                payment_cell.fill = CASH_DARK_FILL
                payment_cell.font = Font(color="FFFFFF", bold=True)
            else:  # credit
                payment_cell.fill = CREDIT_DARK_FILL
                payment_cell.font = Font(color="FFFFFF", bold=True)

        username = cast(str, p.created_by.username) if p.created_by else "—"
        ws.cell(row=row, column=6, value=username)

        row += 1

    _auto_width(ws)



def _create_staff_sheet(
    wb: Workbook,
    sessions: list[Session],
    staff: list[User],
    report_date: dt.date,
):
    """Create sheet with staff working hours and salary calculations."""
    ws = wb.create_sheet(title="Зарплаты персонала")

    # Headers
    headers = ["Сотрудник", "Роль", "Часов", "Ставка/час", "Зарплата"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, len(headers))

    row = 2
    total_salary = 0

    for person in staff:
        username = cast(str, person.username)
        role = cast(str, person.role)
        hourly_rate = int(cast(int, person.hourly_rate)) if person.hourly_rate else 0

        # Calculate hours based on role
        if role == "dealer":
            hours = _calculate_dealer_hours(sessions, int(cast(int, person.id)))
        else:  # waiter
            hours = _calculate_waiter_hours(sessions, int(cast(int, person.id)))

        if hours == 0:
            continue  # Skip staff with no hours

        salary = round(hours * hourly_rate)
        total_salary += salary

        ws.cell(row=row, column=1, value=username)
        ws.cell(row=row, column=2, value="Дилер" if role == "dealer" else "Официант")
        ws.cell(row=row, column=3, value=round(hours, 2))
        ws.cell(row=row, column=4, value=hourly_rate)
        ws.cell(row=row, column=5, value=salary)

        row += 1

    # Total row
    row += 1
    ws.cell(row=row, column=4, value="ИТОГО:")
    ws.cell(row=row, column=4).font = Font(bold=True)
    ws.cell(row=row, column=5, value=total_salary)
    ws.cell(row=row, column=5).font = Font(bold=True)
    ws.cell(row=row, column=5).fill = MONEY_NEGATIVE_FILL  # Salary is an expense

    _auto_width(ws)


def _create_balance_adjustments_sheet(
    wb: Workbook,
    balance_adjustments: list[CasinoBalanceAdjustment],
    report_date: dt.date,
):
    """Create sheet with balance adjustments for the working day."""
    ws = wb.create_sheet(title="Корректировки баланса")

    # Headers
    headers = ["Время", "Тип", "Сумма", "Комментарий", "Создал"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, len(headers))

    if not balance_adjustments:
        ws.cell(row=2, column=1, value="Нет корректировок за выбранную дату")
        ws.cell(row=2, column=1).font = Font(italic=True)
        _auto_width(ws)
        return

    row = 2
    total_profit = 0
    total_expense = 0

    for adj in balance_adjustments:
        time_str = cast(dt.datetime, adj.created_at).strftime("%H:%M:%S")
        amount = int(cast(int, adj.amount))
        
        # Determine type (income/expense)
        adj_type = "Доход" if amount > 0 else "Расход"
        
        ws.cell(row=row, column=1, value=time_str)
        ws.cell(row=row, column=2, value=adj_type)
        
        amount_cell = ws.cell(row=row, column=3, value=amount)
        if amount > 0:
            amount_cell.fill = MONEY_POSITIVE_FILL
            total_profit += amount
        else:
            amount_cell.fill = MONEY_NEGATIVE_FILL
            total_expense += abs(amount)
        
        ws.cell(row=row, column=4, value=cast(str, adj.comment))
        
        username = cast(str, adj.created_by.username) if adj.created_by else "—"
        ws.cell(row=row, column=5, value=username)
        
        row += 1

    # Totals row
    row += 1
    ws.cell(row=row, column=4, value="ИТОГО доходы:")
    ws.cell(row=row, column=4).font = Font(bold=True)
    ws.cell(row=row, column=5, value=total_profit)
    ws.cell(row=row, column=5).font = Font(bold=True)
    ws.cell(row=row, column=5).fill = MONEY_POSITIVE_FILL
    row += 1
    
    ws.cell(row=row, column=4, value="ИТОГО расходы:")
    ws.cell(row=row, column=4).font = Font(bold=True)
    ws.cell(row=row, column=5, value=total_expense)
    ws.cell(row=row, column=5).font = Font(bold=True)
    ws.cell(row=row, column=5).fill = MONEY_NEGATIVE_FILL

    _auto_width(ws)


def _create_summary_sheet(
    wb: Workbook,
    sessions: list[Session],
    seats_by_session: dict[str, list[Seat]],
    purchases: list[ChipPurchase],
    staff: list[User],
    balance_adjustments: list[CasinoBalanceAdjustment],
    report_date: dt.date,
    is_table_admin: bool = False,
):
    """Create summary sheet with profit/expense overview."""
    ws = wb.create_sheet(title="Итоги дня")

    # Calculate totals
    total_chip_income_cash = 0  # Cash buyins (positive only)
    total_chip_cashout = 0  # Cash cashouts (absolute value, negative amounts)
    total_chip_income_credit = 0  # Credit buyins (informational only, NOT in rake calculation)

    for p in purchases:
        amount = int(cast(int, p.amount))
        if amount > 0:  # Buyin
            if p.payment_type == "credit":
                total_chip_income_credit += amount  # Track for info only
            else:
                total_chip_income_cash += amount
        elif amount < 0:  # Cashout (negative amount = expense)
            # Cashouts are always cash payments back to players
            total_chip_cashout += abs(amount)  # Track cashouts separately

    # Calculate balance adjustments (only if not table_admin)
    total_balance_adjustments_profit = 0
    total_balance_adjustments_expense = 0
    if not is_table_admin:
        for adj in balance_adjustments:
            amount = int(cast(int, adj.amount))
            if amount > 0:
                total_balance_adjustments_profit += amount
            else:
                total_balance_adjustments_expense += abs(amount)

    # Calculate staff salary (only if not table_admin)
    total_salary = 0
    if not is_table_admin:
        for person in staff:
            role = cast(str, person.role)
            hourly_rate = int(cast(int, person.hourly_rate)) if person.hourly_rate else 0

            if role == "dealer":
                hours = _calculate_dealer_hours(sessions, int(cast(int, person.id)))
            else:
                hours = _calculate_waiter_hours(sessions, int(cast(int, person.id)))

            total_salary += round(hours * hourly_rate)

    # Calculate net per-seat totals (what players ended with)
    total_player_balance = 0
    for sid, seats in seats_by_session.items():
        for seat in seats:
            total_player_balance += int(cast(int, seat.total))

    # Gross rake ("грязный") = sum of manually entered rake entries from dealer assignments
    gross_rake = 0
    for s in sessions:
        for assignment in s.dealer_assignments:
            for entry in (assignment.rake_entries or []):
                gross_rake += int(cast(int, entry.amount))

    # Net result = gross rake - salaries + balance adjustments (profit/expense)
    net_result = gross_rake - total_salary + total_balance_adjustments_profit - total_balance_adjustments_expense

    # Write summary
    row = 1

    ws.cell(row=row, column=1, value=f"Отчёт за {report_date.isoformat()}")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    row += 2

    # Income section
    ws.cell(row=row, column=1, value="ДОХОДЫ")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=1).fill = MONEY_POSITIVE_FILL
    row += 1

    ws.cell(row=row, column=1, value="Рейк (грязный):")
    ws.cell(row=row, column=2, value=gross_rake)
    ws.cell(row=row, column=2).fill = MONEY_POSITIVE_FILL
    row += 1

    ws.cell(row=row, column=1, value="Покупка фишек (наличные):")
    ws.cell(row=row, column=2, value=total_chip_income_cash)
    ws.cell(row=row, column=2).fill = MONEY_POSITIVE_FILL
    row += 1

    # Only show balance adjustments income for superadmin
    if not is_table_admin:
        ws.cell(row=row, column=1, value="Корректировки баланса (доход):")
        ws.cell(row=row, column=2, value=total_balance_adjustments_profit)
        ws.cell(row=row, column=2).fill = MONEY_POSITIVE_FILL
        row += 1
    row += 1

    # Expense section
    ws.cell(row=row, column=1, value="РАСХОДЫ")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=1).fill = MONEY_NEGATIVE_FILL
    row += 1

    # Only show salaries for superadmin
    if not is_table_admin:
        ws.cell(row=row, column=1, value="Зарплаты персонала:")
        ws.cell(row=row, column=2, value=total_salary)
        ws.cell(row=row, column=2).fill = MONEY_NEGATIVE_FILL
        row += 1

    # Only show balance adjustments expense for superadmin
    if not is_table_admin:
        ws.cell(row=row, column=1, value="Корректировки баланса (расход):")
        ws.cell(row=row, column=2, value=total_balance_adjustments_expense)
        ws.cell(row=row, column=2).fill = MONEY_NEGATIVE_FILL
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="ИТОГО ЗА ДЕНЬ:")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    cell = ws.cell(row=row, column=2, value=net_result)
    cell.font = Font(bold=True, size=12)
    if net_result >= 0:
        cell.fill = MONEY_POSITIVE_FILL
    else:
        cell.fill = MONEY_NEGATIVE_FILL
    row += 2

    # Additional info
    ws.cell(row=row, column=1, value="Справочно:")
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1

    ws.cell(row=row, column=1, value="Баланс игроков (остаток фишек):")
    ws.cell(row=row, column=2, value=total_player_balance)
    row += 1

    ws.cell(row=row, column=1, value="Количество сессий:")
    ws.cell(row=row, column=2, value=len(sessions))
    row += 1

    open_sessions = len([s for s in sessions if s.status == "open"])
    ws.cell(row=row, column=1, value="Открытых сессий:")
    ws.cell(row=row, column=2, value=open_sessions)

    _auto_width(ws)